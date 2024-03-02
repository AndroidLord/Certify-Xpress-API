from fastapi import APIRouter, Query, status, UploadFile, File, HTTPException
import config
import os
import openpyxl
from datetime import datetime
from fastapi.responses import JSONResponse, FileResponse
import s3config
import schema
from util import replace_text_in_docx, convert_to_pdf
import emailconfig
import time

router = APIRouter(tags=["Certification"])

# Routes for Certification

@router.post("/c/", status_code=status.HTTP_201_CREATED)
async def create_certification(certificate: schema.CertificateIn,
                               template: str = Query(..., title="Template", 
                                                     description="Select a template",
                                                     enum=config.template_options)):

    certificate_template = config.CERTIFICATE_TEMPLATE_PATH + f'{template}'
    print(f"Selected template: {certificate_template}")

    doc_output_path = os.path.join(config.DOC_OUTPUT_PATH, f"{certificate.name}.docx")
    pdf_output_path = os.path.join(config.PDF_OUTPUT_PATH, f"{certificate.name}.pdf")

    await replace_text_in_docx(
        template_path=certificate_template,
        certificate=certificate,
        output_path=doc_output_path
    )
    await convert_to_pdf(doc_output_path, pdf_output_path)

    print(f"PDF saved to {pdf_output_path}")

    max_attempts = 10
    curr_attempt = 0
    while not os.path.exists(pdf_output_path):
        curr_attempt +=1
        if curr_attempt > max_attempts:
            print("PDF file could not be created within the specified time.")
            return
        print("Waiting for PDF file to be created...")
        time.sleep(1) 

    if certificate.upload_to_cloud:
        url = await upload_Cert_to_Cloud(certificate.name, pdf_output_path)

    if certificate.send_mail:
        print(f"Sending mail to {certificate.mail}...")
        await emailconfig.send_email([certificate.mail],[pdf_output_path])

    return {
        "name": certificate.name,
        "certification_type": certificate.certification_type,
        "achivement_type": certificate.achivement_type,
        "mentor_name": certificate.mentor_name,
        "mentor_type": certificate.mentor_type,
        "insititue_name": certificate.insititue_name,
        "course_name": certificate.course_name,
        "mail": certificate.mail,
        "url": url if certificate.upload_to_cloud else None,
        "download_pdf_path": pdf_output_path,
        "send_mail": certificate.send_mail,
        "upload_to_cloud": certificate.upload_to_cloud,
        "cloud_url": url if certificate.upload_to_cloud else None
    }
        

@router.post("/c/excel/1", status_code=status.HTTP_201_CREATED)
async def create_certification_from_excel(
        username: str,
        upload_to_cloud: bool = False,
        send_mail: bool = True,
        files: UploadFile = File(..., title="Excel file", description="Upload an excel file"),
        template: str = Query(..., title="Template", description="Select a template", enum=config.template_options)
):
    
        # Create a unique name for the uploaded file
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
        file_name = f"{username}_Excel_{files.filename}_{timestamp}.xlsx"
        excel_file_path = os.path.join(config.EXCEL_ROOT_PATH, file_name)

    
        # Save the uploaded Excel file
        with open(excel_file_path, "wb") as buffer:
            buffer.write(await files.read())

        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        certificate_template = os.path.join(config.CERTIFICATE_TEMPLATE_PATH, template)
        print(f"Selected template: {certificate_template}")

        name = username

        # Update doc_path to a unique name with timestamp
        # Update doc_output_path to a unique name with timestamp
        doc_output_path = os.path.join(config.DOC_OUTPUT_PATH, f"{name}-{timestamp}.docx")

        excel_heading_dict = {
            "Name": "name",
            "Course Name": "course_name",
            "Certificate Type": "certification_type",
            "Achievement Type": "achivement_type",
            "Institute Name": "insititue_name",
            "Mentor Name": "mentor_name",
            "Mentor Type": "mentor_type",
            "mail" : "mail",
            "Director Name": "director_name",
            "Director Type": "director_type"

        }

        match_excel_heading_dict = {}
        heading_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        
        # iterate through the rows in the excel file using i and val
        for i, val in enumerate(heading_row):
            if val in excel_heading_dict:
                match_excel_heading_dict[i] = excel_heading_dict[val]
        
        certificate_links = []
        local_certificates_path = []
        mails = []

        # iterate through the rows in the excel file, but the starting row will be 2     
        for row in sheet.iter_rows(min_row=2, values_only=True):

            certificate_data = {}
            # iterate through the columns in the current row
            for col_index, col_value in enumerate(row):
                if col_index in match_excel_heading_dict:
                    field_name = match_excel_heading_dict[col_index]
                    print(f"{field_name}: {col_value}")
                    certificate_data[field_name] = col_value

            if certificate_data['name'] is None:
                # print("No name found, skipping...")
                break

            if certificate_data['mail'] is not None:
                mails.append(certificate_data['mail'])

            await replace_text_in_docx(
                template_path=certificate_template,
                certificate=schema.CertificateIn(**certificate_data),
                output_path=doc_output_path
            )

            pdf_output_path = os.path.join(config.PDF_OUTPUT_PATH, name, f"{certificate_data['name']}.pdf")

            if not os.path.exists(os.path.join(config.PDF_OUTPUT_PATH, name)):
                os.makedirs(os.path.join(config.PDF_OUTPUT_PATH, name))

            try:
                await convert_to_pdf(doc_output_path, pdf_output_path)
                
                max_attempts = 10                                                                                             
                curr_attempt = 0

                while not os.path.exists(pdf_output_path):                                                  
                    curr_attempt +=1                                                                                                
                    if curr_attempt > max_attempts:                       
                        print("PDF file could not be created within the specified time.")                                           
                    print("Waiting for PDF file to be created...")                                                                time.sleep(1)

                if upload_to_cloud:
                    s3_client = s3config.S3Config()
                    url = s3_client.upload_certificate(
                    pdf_output_path,
                    client_name=name,
                    certificate_name=certificate_data['name'])
                
                    certificate_links.append(url if url else None)

                local_certificates_path.append(pdf_output_path if pdf_output_path else None)

                # print(f"Certificate saved & uploaded to {url}")

                print(f"PDF saved to {pdf_output_path}")
            except Exception as e:
                print(f"An error occurred during PDF conversion: {e}")

            print(f"Certificate generated for {certificate_data['name']}")
            print(f"PDF saved to {pdf_output_path}")

        wb.close()
        print("Certificates generated successfully!")

        # remove template docx file
        print(f"Removing {doc_output_path}...")
        os.remove(doc_output_path)

        if upload_to_cloud:
            # Add the certificate links to the Excel file
            excel_file_path = await adding_certificate_link_to_excel(certificate_links, excel_file_path)

        # Send mail to the users
        if send_mail:
            print(f"Sending mail to {mails}...")
            await emailconfig.send_email(mails, local_certificates_path)

        # Removing the local certificates
        for local_cert_path in local_certificates_path:
            print(f"Removing {local_cert_path}...")
            os.remove(local_cert_path)

        return {
            "status": status.HTTP_201_CREATED,
            "excel_file_path": excel_file_path,
            "certificate_links": certificate_links
        }


@router.get("/download/c")
async def download_certificate(pdf_file_path: str, certificate_name: str, course_name: str):
    # Check if the file exists
    if not os.path.exists(pdf_file_path):
        raise HTTPException(status_code=404, detail="Certificate PDF not found")
    
    # Return the file as response
    return FileResponse(pdf_file_path, filename=f"{certificate_name}_{course_name}.pdf", media_type='application/pdf')

    

@router.delete("/delete/c")
async def delete_certificate(pdf_file_path: str):
    # Check if the file exists
    if not os.path.exists(pdf_file_path):
        raise HTTPException(status_code=404, detail="Certificate PDF not found")
    
    # Delete the file
    os.remove(pdf_file_path)
    
    return {"message": "Certificate deleted successfully!"}


# write routes for downloading and deleteing the excel file

@router.get("/download/excel")
async def download_excel(excel_file_path: str):
    # Check if the file exists
    if not os.path.exists(excel_file_path):
        raise HTTPException(status_code=404, detail="Excel file not found")
    
    # Return the file as response
    return FileResponse(excel_file_path, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@router.delete("/delete/excel")
async def delete_excel(excel_file_path: str):
    # Check if the file exists
    if not os.path.exists(excel_file_path):
        raise HTTPException(status_code=404, detail="Excel file not found")
    
    # Delete the file
    os.remove(excel_file_path)
    
    return {"message": "Excel file deleted successfully!"}


# Routes for Certificate Templates

@router.get("/templates")
async def add_certificate_templates(uploaded_Certificate: UploadFile = File(... , title="Upload Certificate", 
                                                                            description="Upload a certificate")):
    
    # Save the uploaded certificate

    certificate_template_path = os.path.join(config.CERTIFICATE_TEMPLATE_PATH, uploaded_Certificate.filename)
    with open(certificate_template_path, "wb") as buffer:
        buffer.write(await uploaded_Certificate.read())

    print(f"Certificate template saved to {certificate_template_path}")


# Auxiliary OR Helper Functions 


async def adding_certificate_link_to_excel(certificate_links, excel_file_path):
   
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    # Add a header for the new column
    sheet.cell(row=1, column=sheet.max_column + 1, value='Certificate Link')

    # Iterate over each row and add the corresponding Certificate Link
    for row_num in range(2, min(sheet.max_row + 1, len(certificate_links) + 2)):  
        # Start from row 2, assuming row 1 is header
        # Adjust index to match the list of links
        certificate_link = certificate_links[row_num - 2] if row_num - 2 < len(certificate_links) else 'Null'
        sheet.cell(row=row_num, column=sheet.max_column, value=certificate_link)


    # Save the updated Excel file
    prev_excel_file_path = excel_file_path
    excel_file_path = os.path.join(config.EXCEL_ROOT_PATH, f"CertifyXpress_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.xlsx")
    wb.save(excel_file_path)
    
    # Remove the previous Excel file
    print(f"Removing {prev_excel_file_path}...")
    os.remove(prev_excel_file_path)

    print(f"Certificate links added to {excel_file_path}")
    wb.close()
    return excel_file_path



async def upload_Cert_to_Cloud(certificate_name, pdf_output_path):
    # Ask user if they want to upload the certificate to S3
        s3_client = s3config.S3Config()
        url = s3_client.upload_certificate(
            pdf_output_path,
            client_name=certificate_name,
            certificate_name=certificate_name)
        print(f"Certificate uploaded to {url}")

        return url



# END
  
    
