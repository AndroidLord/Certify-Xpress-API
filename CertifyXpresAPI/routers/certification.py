from fastapi import APIRouter, Query, status, UploadFile, File, HTTPException
import config
import os
import openpyxl
from datetime import datetime
from fastapi.responses import JSONResponse, FileResponse
import s3config
import schema
from util import replace_text_in_docx, convert_to_pdf
import emailconfig

router = APIRouter(tags=["Certification"])

@router.post("/c/1", status_code=status.HTTP_201_CREATED)
async def create_certification(certificate: schema.CertificateIn1,
                               template: str = Query(..., title="Template", description="Select a template",
                                                    enum=config.template_options)):

    certificate_template = config.CERTIFICATE_TEMPLATE_PATH + f'{template}'
    print(f"Selected template: {certificate_template}")

    doc_output_path = os.path.join(config.DOC_OUTPUT_PATH, f"{certificate.name}.docx")
    pdf_output_path = os.path.join(config.PDF_OUTPUT_PATH, f"{certificate.name}.pdf")

    await replace_text_in_docx(
        template_path=certificate_template,
        certificate=certificate,
        output_path=doc_output_path
    )
    await convert_to_pdf(doc_output_path, pdf_output_path)

    # Ask user if they want to upload the certificate to S3
    s3_client = s3config.S3Config()
    url = s3_client.upload_certificate(
        pdf_output_path,
        client_name=certificate.name,
        certificate_name=certificate.name)
    print(f"Certificate uploaded to {url}")

    print(f"PDF saved to {pdf_output_path}")
    print("Certificate generated successfully!")

    return {
            "message": "Certificate generated successfully!",
            "download path": pdf_output_path,
            "Online url": url
        }  
    
    
@router.post("/c/2", status_code=status.HTTP_201_CREATED)
async def create_certification(certificate: schema.CertificateIn2,
                               template: str = Query(..., title="Template", description="Select a template",
                                                    enum=config.template_options)):

    certificate_template = config.CERTIFICATE_TEMPLATE_PATH + f'{template}'
    print(f"Selected template: {certificate_template}")

    doc_output_path = fr'{config.DOC_OUTPUT_PATH}/{certificate.name}.docx'
    pdf_output_path = fr'{config.PDF_OUTPUT_PATH}/{certificate.name}.pdf'

    await replace_text_in_docx(
        template_path=certificate_template,
        certificate=certificate,
        output_path=doc_output_path
    )
    await convert_to_pdf(doc_output_path, pdf_output_path)

    # Ask user if they want to upload the certificate to S3

    s3_client = s3config.S3Config()
    url = s3_client.upload_certificate(
        pdf_output_path,
        client_name=certificate.name,
        certificate_name=certificate.name)
    print(f"Certificate uploaded to {url}")

    print(f"PDF saved to {pdf_output_path}")
    print("Certificate generated successfully!")

    return {
        "message": "Certificate generated successfully!",
        "local path": pdf_output_path,
        "Online url": url
    }


@router.post("/c/excel/1", status_code=status.HTTP_201_CREATED)
async def create_certification_from_excel(
        files: UploadFile = File(..., title="Excel file", description="Upload an excel file"),
        template: str = Query(..., title="Template", description="Select a template", enum=config.template_options)
):
    
        # Create a unique name for the uploaded file
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S%f")
        file_name = f"{timestamp}_{files.filename}"
        excel_file_path = os.path.join(config.EXCEL_ROOT_PATH, file_name)

    
        # Save the uploaded Excel file
        with open(excel_file_path, "wb") as buffer:
            buffer.write(await files.read())

        wb = openpyxl.load_workbook(excel_file_path)
        sheet = wb.active

        certificate_template = os.path.join(config.CERTIFICATE_TEMPLATE_PATH, template)
        print(f"Selected template: {certificate_template}")

        name = "shubhamXYZ"

        # Update doc_path to a unique name with timestamp
        # Update doc_output_path to a unique name with timestamp
        doc_output_path = os.path.join(config.DOC_OUTPUT_PATH, f"{name}-{timestamp}.docx")

        excel_heading_dict = {
            "Name": "name",
            "Course Name": "course_name",
            "Certificate Type": "certification_type",
            "Achievement Type": "achivement_type",
            "Institute Name": "insititue_name",
            "Mentor Name": "mentor_name",
            "Mentor Type": "mentor_type",
            "mail" : "mail"
        }

        match_excel_heading_dict = {}
        heading_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        
        # iterate through the rows in the excel file using i and val
        for i, val in enumerate(heading_row):
            if val in excel_heading_dict:
                match_excel_heading_dict[i] = excel_heading_dict[val]
        
        certificate_links = []
        local_certificates_path = []
        mails = []

        # iterate through the rows in the excel file, but the starting row will be 2     
        for row in sheet.iter_rows(min_row=2, values_only=True):

            certificate_data = {}
            # iterate through the columns in the current row
            for col_index, col_value in enumerate(row):
                if col_index in match_excel_heading_dict:
                    field_name = match_excel_heading_dict[col_index]
                    print(f"{field_name}: {col_value}")
                    certificate_data[field_name] = col_value

            if certificate_data['name'] is None:
                # print("No name found, skipping...")
                break
            if certificate_data['mail'] is not None:
                mails.append(certificate_data['mail'])

            await replace_text_in_docx(
                template_path=certificate_template,
                certificate=schema.CertificateIn1(**certificate_data),
                output_path=doc_output_path
            )

            
                

            pdf_output_path = os.path.join(config.PDF_OUTPUT_PATH, name, f"{certificate_data['name']}.pdf")

            if not os.path.exists(os.path.join(config.PDF_OUTPUT_PATH, name)):
                os.makedirs(os.path.join(config.PDF_OUTPUT_PATH, name))

            try:
                await convert_to_pdf(doc_output_path, pdf_output_path)

                s3_client = s3config.S3Config()
                url = s3_client.upload_certificate(
                    pdf_output_path,
                    client_name=name,
                    certificate_name=certificate_data['name'])
                
                certificate_links.append(url if url else None)
                local_certificates_path.append(pdf_output_path if pdf_output_path else None)

                # print(f"Certificate saved & uploaded to {url}")

                print(f"PDF saved to {pdf_output_path}")
            except Exception as e:
                print(f"An error occurred during PDF conversion: {e}")

            print(f"Certificate generated for {certificate_data['name']}")
            print(f"PDF saved to {pdf_output_path}")

        wb.close()
        print("Certificates generated successfully!")


        
        # Add the certificate links to the Excel file
        await adding_certificate_link_to_excel(certificate_links, excel_file_path)

        # Send mail to the users
        await emailconfig.send_email(mails, local_certificates_path)


        return FileResponse(excel_file_path, filename=f"{file_name}", media_type='application/*')

# def send_certificate(pdf_file_path: str):
    
    # # Check if the file exists
    # if not os.path.exists(pdf_file_path):
    #     raise HTTPException(status_code=404, detail="Certificate PDF not found")
    
    # # Return the file as response
    # return FileResponse(pdf_file_path,filename='certificate', media_type='application/pdf')


@router.get("/download/")
async def download_certificate(pdf_file_path: str, certificate_name: str, course_name: str):
    # Check if the file exists
    if not os.path.exists(pdf_file_path):
        raise HTTPException(status_code=404, detail="Certificate PDF not found")
    
    # Return the file as response
    return FileResponse(pdf_file_path, filename=f"{certificate_name}_{course_name}.pdf", media_type='application/pdf')

    
@router.delete("/delete/")
async def delete_certificate(pdf_file_path: str):
    # Check if the file exists
    if not os.path.exists(pdf_file_path):
        raise HTTPException(status_code=404, detail="Certificate PDF not found")
    
    # Delete the file
    os.remove(pdf_file_path)
    
    return {"message": "Certificate deleted successfully!"}


async def adding_certificate_link_to_excel(certificate_links, excel_file_path):
   
    wb = openpyxl.load_workbook(excel_file_path)
    sheet = wb.active

    # Add a header for the new column
    sheet.cell(row=1, column=sheet.max_column + 1, value='Certificate Link')

    # Iterate over each row and add the corresponding Certificate Link
    for row_num in range(2, min(sheet.max_row + 1, len(certificate_links) + 2)):  
        # Start from row 2, assuming row 1 is header
        # Adjust index to match the list of links
        certificate_link = certificate_links[row_num - 2] if row_num - 2 < len(certificate_links) else 'Null'
        sheet.cell(row=row_num, column=sheet.max_column, value=certificate_link)


    # Save the updated Excel file
    wb.save('certifyXpress_Shubham_' + datetime.now().strftime("%Y%m%d-%H%M%S") + '.xlsx')
    print(f"Certificate links added to {excel_file_path}")
    wb.close()



  
    